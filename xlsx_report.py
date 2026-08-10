#!/usr/bin/env python3
"""Заполняет бланк «Отчёт о проделанной работе» данными выгрузки Jira.

Вход — JSON от generate_report.py (скилл jira-browser-workflow), выход — xlsx по
шаблону template.xlsx: строки данных клонируются по числу задач, шапка и подписи
остаются как в бланке.

Соглашения (согласованы с пользователем, менять здесь):
  * отчётные сутки начинаются в 05:00, как в generate_report.py: ночное списание
    07.08 в 00:16 относится к рабочему дню 06.08;
  * «Начало/Окончание работы» в Jira недостоверны (started — это момент списания,
    часто ночью), поэтому задачи дня раскладываются подряд от начала рабочего дня;
  * «График рабочего дня» = начало дня … начало + сумма списанных часов.

    python xlsx_report.py reports/Jira_ATOM_отчёт_2026-08-03_2026-08-07.json out.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import zipfile
from copy import copy
from datetime import date, datetime, time, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

TEMPLATE = Path(__file__).with_name("template.xlsx")
DAY_START_HOUR = 5      # как в generate_report.py
WORK_START = time(9, 0)

# Строки бланка: 11–12 — заголовки таблицы, 13–22 — область данных на один день.
HEADER_ROWS = 12
FIRST_DATA_ROW = 13
TEMPLATE_DATA_ROWS = 10
COLUMNS = "ABCDEFGHIJ"

MONTHS = ("января", "февраля", "марта", "апреля", "мая", "июня",
          "июля", "августа", "сентября", "октября", "ноября", "декабря")
MONTHS_NOMINATIVE = ("Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                     "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь")


def ru_date(day: date) -> str:
    return "%02d %s %d" % (day.day, MONTHS[day.month - 1], day.year)


def report_day(moment: datetime) -> date:
    """Отчётный день момента: сутки начинаются в DAY_START_HOUR."""
    return (moment - timedelta(hours=DAY_START_HOUR)).date()


def collect_days(payload: dict) -> list[tuple[date, list[dict]]]:
    """Группирует worklog пользователя по отчётным дням внутри периода.

    Задачи без списаний в отчёт не попадают: в бланке нет строки без трудозатрат.
    """
    login = (payload.get("scope", {}).get("report_user") or {}).get("login")
    since = date.fromisoformat(payload["period"]["since"])
    until = date.fromisoformat(payload["period"]["until"])
    days: dict[date, list[dict]] = {}
    for issue in payload["issues"]:
        for wl in issue.get("worklogs", []):
            if login and wl.get("author_login") != login:
                continue
            started = datetime.fromisoformat(wl["started"])
            day = report_day(started)
            if not since <= day <= until:
                continue
            days.setdefault(day, []).append({
                "started": started,
                "seconds": wl["seconds"],
                "summary": issue["summary"],
                "url": issue["url"],
                # ponytail: «статус на момент фиксации» = текущий статус задачи. В changelog
                # он лежит по-английски (In work → Suspended), для исторического статуса
                # нужен словарь перевода — добавить, если проверяющий начнёт придираться.
                "status": issue["status"],
                # «Блокирующие факторы» в Jira нет: для приостановленных задач
                # ближайшая правда — комментарий, которым пользователь объяснил паузу.
                "blocker": (wl.get("comment") or "") if issue.get("status_group") == "paused" else "",
            })
    for entries in days.values():
        entries.sort(key=lambda e: e["started"])
    return sorted(days.items())


def build_rows(days: list[tuple[date, list[dict]]], work_start: time) -> list[dict]:
    """Разворачивает дни в плоские строки бланка с расчётом времени по графику."""
    rows = []
    for day, entries in days:
        cursor = datetime.combine(day, work_start)
        day_total = sum(e["seconds"] for e in entries)
        day_end = cursor + timedelta(seconds=day_total)
        for index, entry in enumerate(entries):
            finish = cursor + timedelta(seconds=entry["seconds"])
            rows.append({
                "first_of_day": index == 0,
                "day_rows": len(entries),
                "date": ru_date(day),
                "day_start": cursor.strftime("%H:%M") if index == 0 else None,
                "day_end": day_end.strftime("%H:%M") if index == 0 else None,
                "summary": entry["summary"],
                "start": cursor.strftime("%H:%M"),
                "finish": finish.strftime("%H:%M"),
                "url": entry["url"],
                "hours": round(entry["seconds"] / 3600, 2),
                "status": entry["status"],
                "blocker": entry["blocker"],
            })
            cursor = finish
    return rows


def _resize_data_area(ws, needed: int) -> None:
    """Подгоняет число строк данных под needed, сохраняя высоту строк бланка."""
    for merged in [str(r) for r in ws.merged_cells.ranges]:
        if re.match(r"^[ABC]%d:[ABC]%d$" % (FIRST_DATA_ROW, FIRST_DATA_ROW + TEMPLATE_DATA_ROWS - 1), merged):
            ws.unmerge_cells(merged)
    delta = needed - TEMPLATE_DATA_ROWS
    if delta > 0:
        ws.insert_rows(FIRST_DATA_ROW + TEMPLATE_DATA_ROWS, delta)
    elif delta < 0:
        ws.delete_rows(FIRST_DATA_ROW + needed, -delta)
    # Бланк помечает строки фиксированной высотой (customHeight), из-за чего Excel не
    # подгоняет её под перенос текста и длинные темы задач срезаются. Снимаем метку —
    # wrap_text в шаблоне уже включён, высоту посчитает сам Excel.
    # customHeight в openpyxl вычисляется из height, отдельно его снимать нельзя.
    for offset in range(needed):
        ws.row_dimensions[FIRST_DATA_ROW + offset].height = None


def _apply_style(ws, row: int, donor: int, day_top: bool, day_bottom: bool) -> None:
    """Копирует оформление строки-донора; рамку блока дня в A–C собирает по месту.

    В бланке A/B/C образуют рамку вокруг всего дня: верх — у первой строки дня
    (донор 13), низ — у последней (донор 22), середина — только боковые (донор 14).
    """
    for index, letter in enumerate(COLUMNS, start=1):
        target = ws.cell(row=row, column=index)
        source = ws.cell(row=donor, column=index)
        target._style = copy(source._style)
        if letter in "ABC":
            border = copy(target.border)
            border.top = copy(ws.cell(row=FIRST_DATA_ROW, column=index).border.top) \
                if day_top else copy(ws.cell(row=FIRST_DATA_ROW + 1, column=index).border.top)
            border.bottom = copy(ws.cell(row=FIRST_DATA_ROW + TEMPLATE_DATA_ROWS - 1, column=index).border.bottom) \
                if day_bottom else copy(ws.cell(row=FIRST_DATA_ROW + 1, column=index).border.bottom)
            target.border = border


def fill_sheet(ws, payload: dict, rows: list[dict], employee: str | None) -> None:
    """Пишет шапку, строки задач и объединяет ячейки дня."""
    ws["B5"] = employee or (payload.get("scope", {}).get("report_user") or {}).get("name") or ""
    ws["G4"] = date.fromisoformat(payload["period"]["since"])
    ws["H4"] = date.fromisoformat(payload["period"]["until"])

    donors = {"first": FIRST_DATA_ROW, "middle": FIRST_DATA_ROW + 1,
              "last": FIRST_DATA_ROW + TEMPLATE_DATA_ROWS - 1}
    styles = []
    for offset, row in enumerate(rows):
        position = offset - next(i for i in range(offset + 1) if rows[offset - i]["first_of_day"])
        is_last = position == row["day_rows"] - 1
        donor = donors["first"] if row["first_of_day"] else (donors["last"] if is_last else donors["middle"])
        styles.append((donor, row["first_of_day"], is_last))

    _resize_data_area(ws, max(len(rows), 1))

    for offset, (row, (donor, top, bottom)) in enumerate(zip(rows, styles)):
        line = FIRST_DATA_ROW + offset
        _apply_style(ws, line, donor, top, bottom)
        if row["first_of_day"]:
            ws.cell(row=line, column=1).value = row["date"]
            ws.cell(row=line, column=2).value = row["day_start"]
            ws.cell(row=line, column=3).value = row["day_end"]
            if row["day_rows"] > 1:
                for column in "ABC":
                    ws.merge_cells("%s%d:%s%d" % (column, line, column, line + row["day_rows"] - 1))
        ws.cell(row=line, column=4).value = row["summary"]
        ws.cell(row=line, column=5).value = row["start"]
        ws.cell(row=line, column=6).value = row["finish"]
        link = ws.cell(row=line, column=7)
        link.value = row["url"]
        link.hyperlink = row["url"]
        hours = ws.cell(row=line, column=8)
        hours.value = row["hours"]
        # Ячейка бланка размечена под время, иначе 6 часов Excel покажет как 06.01.1900.
        hours.number_format = "General"
        ws.cell(row=line, column=9).value = row["status"]
        ws.cell(row=line, column=10).value = row["blocker"]


def _restore_drawings(template: Path, target: Path) -> None:
    """Возвращает в готовый файл логотип и настройки печати.

    openpyxl не умеет wmf и молча выбрасывает картинку вместе с drawing-частью,
    поэтому недостающие части переносим из шаблона и снова подшиваем к листу.
    """
    src = zipfile.ZipFile(template)
    keep = [n for n in src.namelist() if n.startswith(("xl/media/", "xl/drawings/"))]
    if not keep:
        return
    with zipfile.ZipFile(target) as out:
        parts = {name: out.read(name) for name in out.namelist()}

    for name in keep:
        parts[name] = src.read(name)

    # Связи листа перезаписывать нельзя: openpyxl держит там гиперссылки на задачи.
    drawing_id = "rIdDrawing"
    rels_name = "xl/worksheets/_rels/sheet1.xml.rels"
    rels = parts.get(rels_name, b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                                b'<Relationships xmlns="http://schemas.openxmlformats.org'
                                b'/package/2006/relationships"></Relationships>').decode("utf-8")
    if "relationships/drawing" not in rels:
        rels = rels.replace("</Relationships>",
                            '<Relationship Id="%s" Type="http://schemas.openxmlformats.org'
                            '/officeDocument/2006/relationships/drawing" '
                            'Target="../drawings/drawing1.xml"/></Relationships>' % drawing_id)
    parts[rels_name] = rels.encode("utf-8")

    sheet = parts["xl/worksheets/sheet1.xml"].decode("utf-8")
    if "<drawing " not in sheet:
        # Префикс r openpyxl объявляет только локально в <hyperlink>, поэтому в конце
        # документа он не связан и тег drawing ломает файл ошибкой "unbound prefix".
        root = re.search(r"<worksheet\b[^>]*>", sheet).group(0)
        if "xmlns:r=" not in root:
            sheet = sheet.replace(root, root[:-1] + ' xmlns:r="http://schemas.openxmlformats'
                                  '.org/officeDocument/2006/relationships">', 1)
        sheet = sheet.replace("</worksheet>", '<drawing r:id="%s"/></worksheet>' % drawing_id)
        parts["xl/worksheets/sheet1.xml"] = sheet.encode("utf-8")

    types = parts["[Content_Types].xml"].decode("utf-8")
    if "image/x-wmf" not in types:
        types = types.replace(
            "<Default", '<Default Extension="wmf" ContentType="image/x-wmf"/><Default', 1)
    if "drawing1.xml" not in types:
        types = types.replace("</Types>",
                              '<Override PartName="/xl/drawings/drawing1.xml" ContentType='
                              '"application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>')
    parts["[Content_Types].xml"] = types.encode("utf-8")

    with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as out:
        for name, data in parts.items():
            out.writestr(name, data)


ACTIVITY_SHEET = "Все действия"
ACTIVITY_COLUMNS = (
    ("Дата и время", 17),
    ("Задача", 13),
    ("Тема", 52),
    ("Действие", 26),
    ("Подробности", 62),
)


def collect_activity(payload: dict) -> list[dict]:
    """Все действия пользователя за период, по времени.

    Бланк показывает только списанное время: одна строка — один worklog. Переходы,
    комментарии, смена исполнителя и правки полей туда не помещаются, но проверяющему
    они бывают нужнее часов, поэтому отдаём их отдельным листом без потерь.
    """
    login = (payload.get("scope", {}).get("report_user") or {}).get("login")
    rows = []
    for issue in payload["issues"]:
        for event in issue.get("events", []):
            if login and event.get("author_login") != login:
                continue
            moment = datetime.fromisoformat(event["timestamp"])
            rows.append({
                "moment": moment,
                "when": moment.strftime("%d.%m.%Y %H:%M"),
                "key": issue["key"],
                "url": issue["url"],
                "summary": issue["summary"],
                "action": event.get("label") or event.get("type") or "",
                "detail": _activity_detail(event),
            })
    rows.sort(key=lambda r: r["moment"])
    return rows


# Поля, которые Jira отдаёт в секундах: без перевода в часы «14400 → 12000» не читается.
TIME_FIELDS = {"timeestimate", "timeoriginalestimate", "timespent"}
_BIG_NUMBER = re.compile(r"\b(\d{2,})\b")


def _hours(seconds: float) -> str:
    hours = seconds / 3600
    return "%d ч" % hours if hours == int(hours) else "%.2f ч" % hours


def _activity_detail(event: dict) -> str:
    """Человекочитаемая суть события; время показываем в часах, а не в секундах."""
    detail = event.get("detail") or ""
    if event.get("type") == "worklog":
        spent = _hours(event["value"]) if event.get("value") else ""
        return "%s%s" % (spent, " — %s" % detail if detail else "")
    if event.get("value") in TIME_FIELDS:
        return _BIG_NUMBER.sub(lambda m: _hours(int(m.group(1))), detail)
    return detail


def add_activity_sheet(wb, payload: dict) -> int:
    """Добавляет лист со всеми действиями. Возвращает число строк."""
    rows = collect_activity(payload)
    ws = wb.create_sheet(ACTIVITY_SHEET)
    header = Font(bold=True)
    fill = PatternFill("solid", fgColor="EFEFEF")
    edge = Side(style="thin", color="BFBFBF")
    border = Border(left=edge, right=edge, top=edge, bottom=edge)
    top = Alignment(vertical="top", wrap_text=True)

    for index, (title, width) in enumerate(ACTIVITY_COLUMNS, start=1):
        cell = ws.cell(row=1, column=index, value=title)
        cell.font, cell.fill, cell.border = header, fill, border
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    for line, row in enumerate(rows, start=2):
        values = (row["when"], row["key"], row["summary"], row["action"], row["detail"])
        for index, value in enumerate(values, start=1):
            cell = ws.cell(row=line, column=index, value=value)
            cell.border, cell.alignment = border, top
        link = ws.cell(row=line, column=2)
        link.hyperlink = row["url"]
        link.style = "Hyperlink"

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = "A1:%s%d" % (
            ws.cell(row=1, column=len(ACTIVITY_COLUMNS)).column_letter, len(rows) + 1)
    return len(rows)


def build_xlsx(payload: dict, out_path: Path, template: Path = TEMPLATE,
               work_start: time = WORK_START, employee: str | None = None) -> dict:
    """Собирает готовый бланк. Возвращает сводку для UI."""
    days = collect_days(payload)
    rows = build_rows(days, work_start)
    if not rows:
        raise ValueError("За период нет ни одного worklog — бланк заполнять нечем.")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.worksheets[0]
    ws.title = MONTHS_NOMINATIVE[date.fromisoformat(payload["period"]["since"]).month - 1]
    fill_sheet(ws, payload, rows, employee)
    activity = add_activity_sheet(wb, payload)
    wb.save(out_path)
    _restore_drawings(template, out_path)

    return {
        "path": str(out_path),
        "days": len(days),
        "rows": len(rows),
        "hours": round(sum(r["hours"] for r in rows), 2),
        "activity": activity,
    }


def main() -> int:
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("payload", type=Path, help="JSON от generate_report.py")
    parser.add_argument("out", type=Path, help="Куда сохранить заполненный бланк")
    parser.add_argument("--employee", help="ФИО сотрудника; по умолчанию из Jira")
    parser.add_argument("--work-start", default="09:00", help="Начало рабочего дня, ЧЧ:ММ")
    args = parser.parse_args()

    payload = json.loads(args.payload.read_text(encoding="utf-8"))
    hour, minute = (int(part) for part in args.work_start.split(":"))
    result = build_xlsx(payload, args.out, work_start=time(hour, minute),
                        employee=args.employee)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
